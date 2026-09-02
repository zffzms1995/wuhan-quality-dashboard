#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""一键更新看板：飞书云表格导出 → 解析 → 生成数据文件 → git 提交

用法:
  python3 update_dashboard.py                     # 完整流程（导出+解析+提交）
  python3 update_dashboard.py --xlsx 文件.xlsx     # 只解析本地 xlsx（飞书导出失败时手动导出用）
  python3 update_dashboard.py --token <文档token>  # 更换主质量表文档
  python3 update_dashboard.py --cuoti-token <文档token>  # 更换错题/培训留底表
  python3 update_dashboard.py --no-commit          # 只生成数据文件，不提交 git
"""
import argparse
import calendar
import json
import os
import re
import shutil
import subprocess
import time
import urllib.error
import urllib.request
import warnings
from datetime import datetime, date, timedelta

warnings.filterwarnings("ignore")

BASE = "https://open.feishu.cn"
FEISHU_DOC_BASE = "https://zhuanspirit.feishu.cn"
REPO = os.path.dirname(os.path.abspath(__file__))
CLAUDE_JSON = os.path.expanduser("~/.claude.json")
DEFAULT_FILE_TOKEN = "CzCEs7FJ9hozKKtjRRJcC4ilnne"
CUOTI_FILE_TOKEN = "XJn6sd0WFhY5SPtDYoLcZcXenyf"
ARCHIVE_DIR = os.path.join(REPO, "archive")


# ---------- 飞书 API ----------

def _http_json(url, method="GET", headers=None, body=None, timeout=120):
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(url, data=data, method=method)
    req.add_header("Content-Type", "application/json")
    for k, v in (headers or {}).items():
        req.add_header(k, v)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        raise SystemExit(f"飞书接口请求失败 (HTTP {e.code}): {e.read().decode()[:300]}")
    except urllib.error.URLError as e:
        raise SystemExit(f"网络请求失败: {e.reason}（请检查网络后重试）")


def get_app_credentials():
    try:
        with open(CLAUDE_JSON, encoding="utf-8") as f:
            cfg = json.load(f)
    except (OSError, json.JSONDecodeError):
        raise SystemExit(f"无法读取 {CLAUDE_JSON}，请确认飞书 MCP 已配置（claude mcp add feishu ...）")
    for name, srv in (cfg.get("mcpServers") or {}).items():
        if "feishu" in name.lower():
            args = srv.get("args") or []
            app_id = app_secret = None
            for i, a in enumerate(args):
                if a == "-a" and i + 1 < len(args):
                    app_id = args[i + 1]
                if a == "-s" and i + 1 < len(args):
                    app_secret = args[i + 1]
            if app_id and app_secret:
                return app_id, app_secret
    raise SystemExit("未在 ~/.claude.json 的 mcpServers 中找到飞书应用凭证（-a/-s）")


def fetch_xlsx(token):
    app_id, app_secret = get_app_credentials()
    print("① 获取飞书应用凭证 →", app_id)
    r = _http_json(f"{BASE}/open-apis/auth/v3/tenant_access_token/internal", "POST",
                   body={"app_id": app_id, "app_secret": app_secret})
    if r.get("code") != 0:
        raise SystemExit(f"获取 tenant_access_token 失败: {r.get('msg', r)}")
    t = r["tenant_access_token"]
    h = {"Authorization": f"Bearer {t}"}

    print("② 创建导出任务 ...")
    r = _http_json(f"{BASE}/open-apis/drive/v1/export_tasks?file_extension=xlsx", "POST", headers=h,
                   body={"file_extension": "xlsx", "token": token, "type": "sheet"})
    if r.get("code") != 0:
        raise SystemExit(f"创建导出任务失败: {r.get('msg', r)}")
    ticket = r["data"]["ticket"]

    print("③ 等待导出完成 ...", end="", flush=True)
    for _ in range(100):
        time.sleep(3)
        print(".", end="", flush=True)
        rr = _http_json(f"{BASE}/open-apis/drive/v1/export_tasks/{ticket}?token={token}", headers=h)
        if rr.get("code") != 0:
            raise SystemExit(f"查询导出任务失败: {rr.get('msg', rr)}")
        res = rr["data"]["result"]
        if res["job_status"] == 0:
            print(" 完成")
            req = urllib.request.Request(
                f"{BASE}/open-apis/drive/v1/medias/{res['file_token']}/download", headers=h)
            with urllib.request.urlopen(req, timeout=600) as resp:
                data = resp.read()
            print(f"④ 下载完成（{res.get('file_name', '')}，{len(data) // 1024} KB）")
            return data
        if res["job_status"] >= 3:
            raise SystemExit(f"飞书导出失败（状态码 {res['job_status']}）: {res.get('job_error_msg', '未知错误')}")
    raise SystemExit("导出超时（超过5分钟），请稍后重试")


# ---------- 解析工具 ----------

def sval(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def num(v):
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str) and v.strip():
        try:
            return float(v.strip().replace(",", ""))
        except ValueError:
            pass
    return 0


def col_of(ws, hdr_row, name):
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=hdr_row, column=c).value
        if v and str(v).strip() == name:
            return c
    raise SystemExit(f"在工作表「{ws.title}」第{hdr_row}行未找到列「{name}」，请确认表格结构未变")


def col_of_after(ws, hdr_row, name, after_col):
    for c in range(after_col + 1, ws.max_column + 1):
        v = ws.cell(row=hdr_row, column=c).value
        if v and str(v).strip() == name:
            return c
    raise SystemExit(f"在工作表「{ws.title}」第{hdr_row}行未找到列「{name}」")


def find_header_row(ws, anchors, max_rows=5):
    for r in range(1, max_rows + 1):
        vals = {str(ws.cell(row=r, column=c).value).strip()
                for c in range(1, ws.max_column + 1)}
        if all(a in vals for a in anchors):
            return r
    return None


def find_sheet(wb, keyword, required=True):
    for ws in wb.worksheets:
        if keyword in ws.title:
            return ws
    if required:
        raise SystemExit(f"云表格中未找到包含「{keyword}」的工作表，请确认文档结构未变")
    return None


# ---------- 各工作表解析 ----------

def _write_optimized_image(data, out_path):
    """原图转 JPEG 并限制最长边 1600，加快国内访问 GitHub Pages 的速度；sips 失败时保留原图"""
    tmp = out_path + ".tmp"
    with open(tmp, "wb") as f:
        f.write(data)
    try:
        r = subprocess.run(
            ["sips", "-s", "format", "jpeg", "-s", "formatOptions", "80",
             "-Z", "1600", tmp, "--out", out_path],
            capture_output=True)
        if r.returncode == 0:
            os.remove(tmp)
        else:
            os.replace(tmp, out_path)
    except OSError:
        os.replace(tmp, out_path)


def parse_mb(ws, images_dir):
    hdr = find_header_row(ws, ["是否触发罚款", "质检码", "审核人"])
    if hdr is None:
        raise SystemExit(f"「{ws.title}」表头未识别，请检查工作表结构")
    cols = {n: col_of(ws, hdr, n) for n in [
        "是否触发罚款", "拆机时间", "质检码", "品牌名称", "型号名称", "功能等级",
        "商户站点名称", "后验结果", "拆机命中埋点操作人", "检出项一级名称",
        "检出项二级名称", "检出项三级名称", "审核人", "是否有差异", "差异检出值", "备注"]}
    rows = []
    for r in range(hdr + 1, ws.max_row + 1):
        date_v = sval(ws.cell(row=r, column=cols["拆机时间"]).value)
        if not date_v:
            continue
        rows.append({
            "_row": r,
            "penalty": sval(ws.cell(row=r, column=cols["是否触发罚款"]).value),
            "date": date_v,
            "code": sval(ws.cell(row=r, column=cols["质检码"]).value),
            "brand": sval(ws.cell(row=r, column=cols["品牌名称"]).value),
            "model": sval(ws.cell(row=r, column=cols["型号名称"]).value),
            "grade": sval(ws.cell(row=r, column=cols["功能等级"]).value),
            "merchant": sval(ws.cell(row=r, column=cols["商户站点名称"]).value),
            "result": sval(ws.cell(row=r, column=cols["后验结果"]).value),
            "inspector": sval(ws.cell(row=r, column=cols["拆机命中埋点操作人"]).value),
            "qaLevel1": sval(ws.cell(row=r, column=cols["检出项一级名称"]).value),
            "qaLevel2": sval(ws.cell(row=r, column=cols["检出项二级名称"]).value),
            "qaCheck": sval(ws.cell(row=r, column=cols["检出项三级名称"]).value),
            "qaReviewer": sval(ws.cell(row=r, column=cols["审核人"]).value),
            "hasDiff": sval(ws.cell(row=r, column=cols["是否有差异"]).value),
            "qaFinding": sval(ws.cell(row=r, column=cols["差异检出值"]).value),
            "notes": sval(ws.cell(row=r, column=cols["备注"]).value),
        })
    # 内嵌图片按锚点行号对应记录
    img_by_row = {}
    for img in getattr(ws, "_images", []):
        try:
            row = img.anchor._from.row + 1
        except Exception:
            continue
        img_by_row.setdefault(row, img)
    written = 0
    for rec in rows:
        img = img_by_row.get(rec["_row"])
        rec["imageFile"] = ""
        if img is not None:
            fn = f"mb_{rec['_row']}.jpg"
            data = img._data() if callable(img._data) else img._data
            _write_optimized_image(data, os.path.join(images_dir, fn))
            rec["imageFile"] = fn
            written += 1
    rows.sort(key=lambda d: (d["date"], d["code"]))
    for i, rec in enumerate(rows, 1):
        rec["id"] = i
        del rec["_row"]
    return rows, written


def parse_phone(ws, images_dir):
    hdr = find_header_row(ws, ["抽检日期", "质检码", "执行问题失误人"])
    if hdr is None:
        raise SystemExit(f"「{ws.title}」表头未识别，请检查工作表结构")
    c = {n: col_of(ws, hdr, n) for n in [
        "抽检日期", "大盘有差异设备量", "大盘客观差异设备量", "总抽检量", "总客观差异量",
        "质检操作人", "大盘抽检量",
        "执行问题失误人", "日期", "质检码", "品牌名称", "型号名称", "一级项名称",
        "二级项名称", "原三级项名称", "新三级项名称", "原操作人名称", "新操作人名称",
        "主客观问题分类", "问题类型", "抽检判定的问题类型", "判轻判重分类",
        "后验&入仓", "商户站点"]}
    # 左侧统计表与右侧记录表都有「站点名称」列，记录表取日期列之后的那个
    c["记录表站点名称"] = col_of_after(ws, hdr, "站点名称", c["日期"])
    img_by_row = {}
    for img in getattr(ws, "_images", []):
        try:
            a = img.anchor
            row = (a._from.row if hasattr(a, "_from") else a.from_.row) + 1
            col = a._from.col if hasattr(a, "_from") else a.from_.col
            img_by_row.setdefault(row, []).append((col, img))
        except Exception:
            continue
    trend, person_sum, diffs = [], {}, []
    code_seen = {}
    for r in range(hdr + 1, ws.max_row + 1):
        # 仓维度（每日一行）
        d1 = ws.cell(row=r, column=c["抽检日期"]).value
        if isinstance(d1, datetime):
            trend.append({
                "date": d1.strftime("%m/%d"),
                "samples": num(ws.cell(row=r, column=c["总抽检量"]).value),
                "diffs": num(ws.cell(row=r, column=c["大盘有差异设备量"]).value),
                "objDiff": num(ws.cell(row=r, column=c["大盘客观差异设备量"]).value),
            })
        # 人维度（抽检量按人汇总，复现看板 SUMIF 公式）
        pname = sval(ws.cell(row=r, column=c["质检操作人"]).value)
        if pname:
            person_sum[pname] = person_sum.get(pname, 0) + num(
                ws.cell(row=r, column=c["大盘抽检量"]).value)
        # 大盘记录表（差异明细）
        d40 = ws.cell(row=r, column=c["日期"]).value
        if isinstance(d40, datetime):
            code = sval(ws.cell(row=r, column=c["质检码"]).value)
            classify = sval(ws.cell(row=r, column=c["主客观问题分类"]).value)
            judgment = sval(ws.cell(row=r, column=c["抽检判定的问题类型"]).value)
            orig = sval(ws.cell(row=r, column=c["原操作人名称"]).value)
            # 复现飞书「执行问题失误人」公式：
            # =IF(AND(COUNTIF(AO$3:AO3,AO3)=1,BL3="客观问题",BN3="执行问题",AR3="武汉库"),BI3,"")
            code_seen[code] = code_seen.get(code, 0) + 1
            first_occur = code_seen[code] == 1
            site = sval(ws.cell(row=r, column=c["记录表站点名称"]).value)
            if first_occur and classify == "客观问题" and judgment == "执行问题" and site == "武汉库":
                error_person = orig
            else:
                error_person = ""
            l2 = sval(ws.cell(row=r, column=c["新三级项名称"]).value) or \
                 sval(ws.cell(row=r, column=c["原三级项名称"]).value) or \
                 sval(ws.cell(row=r, column=c["二级项名称"]).value)
            imgs = []
            for _col, img in sorted(img_by_row.get(r, [])):
                fname = f"phone_{r}_{len(imgs) + 1}.jpg"
                _write_optimized_image(img._data(), os.path.join(images_dir, fname))
                imgs.append(fname)
            diffs.append({
                "date": d40.strftime("%Y-%m-%d"),
                "code": code,
                "imgs": imgs,
                "cat": "手机",
                "brand": sval(ws.cell(row=r, column=c["品牌名称"]).value),
                "model": sval(ws.cell(row=r, column=c["型号名称"]).value),
                "level1": sval(ws.cell(row=r, column=c["一级项名称"]).value),
                "level2": l2,
                "classify": classify,
                "problemType": sval(ws.cell(row=r, column=c["问题类型"]).value),
                "qaJudgment": judgment,
                "severity": sval(ws.cell(row=r, column=c["判轻判重分类"]).value),
                "stage": sval(ws.cell(row=r, column=c["后验&入仓"]).value),
                "merchant": sval(ws.cell(row=r, column=c["商户站点"]).value),
                "errorPerson": error_person,
                "origInspector": orig,
                "newInspector": sval(ws.cell(row=r, column=c["新操作人名称"]).value),
            })
    error_count = {}
    for d in diffs:
        if d["errorPerson"]:
            error_count[d["errorPerson"]] = error_count.get(d["errorPerson"], 0) + 1
    diffs.sort(key=lambda d: (d["date"], d["code"]))
    return trend, person_sum, error_count, diffs


def parse_fourcat(ws, images_dir):
    hdr = find_header_row(ws, ["抽检总量", "执行问题失误人", "质检人"])
    if hdr is None:
        raise SystemExit(f"「{ws.title}」表头未识别，请检查工作表结构")
    c = {n: col_of(ws, hdr, n) for n in [
        "日期", "品类", "原操作人", "抽检总量", "差异数量", "客观差异数量",
        "执行问题失误人", "质检码", "站点", "型号", "质检人", "问题分类"]}
    img_by_row = {}
    for img in getattr(ws, "_images", []):
        try:
            a = img.anchor
            row = (a._from.row if hasattr(a, "_from") else a.from_.row) + 1
            col = a._from.col if hasattr(a, "_from") else a.from_.col
            img_by_row.setdefault(row, []).append((col, img))
        except Exception:
            continue
    # 无表头列按与锚定列的相对位置定位（与历史模板一致）
    c["问题类型"] = c["问题分类"] + 1
    c["抽检判定"] = c["问题分类"] + 2
    c["记录表日期"] = col_of_after(ws, hdr, "日期", c["执行问题失误人"])
    # 记录表与左侧统计表存在同名列（品类/站点/角色），记录表锚点须在记录表日期之后查找
    c["记录表品类"] = col_of_after(ws, hdr, "品类", c["记录表日期"])
    c["记录表站点"] = col_of_after(ws, hdr, "站点", c["记录表日期"])
    c["记录表品牌"] = c["记录表品类"] + 1
    c["记录表型号"] = col_of_after(ws, hdr, "型号", c["记录表日期"])
    c["记录表质检人"] = col_of_after(ws, hdr, "质检人", c["记录表日期"])
    # 「角色」列在表头无标签（空表头），按与质检人列的相对位置定位
    c["记录表角色"] = c["记录表质检人"] + 1
    c["记录表判定方式"] = c["记录表型号"] + 1
    c["记录表结果"] = c["记录表型号"] + 2
    c["记录表一级项"] = c["记录表质检人"] + 3
    c["记录表二级项"] = c["记录表质检人"] + 4
    c["记录表明细"] = c["记录表质检人"] + 6

    daily, daily_cat, person_cat, records = {}, {}, {}, []
    code_seen = {}
    for r in range(hdr + 1, ws.max_row + 1):
        d1 = ws.cell(row=r, column=c["日期"]).value
        if isinstance(d1, datetime):
            ds = d1.strftime("%Y-%m-%d")
            cat = sval(ws.cell(row=r, column=c["品类"]).value)
            op = sval(ws.cell(row=r, column=c["原操作人"]).value)
            samples = num(ws.cell(row=r, column=c["抽检总量"]).value)
            diffs_n = num(ws.cell(row=r, column=c["差异数量"]).value)
            daily[ds] = daily.get(ds, [0, 0])
            daily[ds][0] += samples
            daily[ds][1] += diffs_n
            daily_cat.setdefault(cat, {})
            daily_cat[cat][ds] = daily_cat[cat].get(ds, [0, 0])
            daily_cat[cat][ds][0] += samples
            daily_cat[cat][ds][1] += diffs_n
            if op:
                person_cat.setdefault((op, cat), 0)
                person_cat[(op, cat)] += samples
        d17 = ws.cell(row=r, column=c["记录表日期"]).value
        if isinstance(d17, datetime):
            code = sval(ws.cell(row=r, column=c["质检码"]).value)
            classify = sval(ws.cell(row=r, column=c["问题分类"]).value)
            qa_type = sval(ws.cell(row=r, column=c["抽检判定"]).value)
            inspector = sval(ws.cell(row=r, column=c["记录表质检人"]).value)
            # 复现飞书「执行问题失误人」公式：
            # =IF(AND(COUNTIF(R$3:R7,R7)=1,AF7="客观问题",AH7="执行问题",Z7="质检",S7="武汉库"),Y7,"")
            code_seen[code] = code_seen.get(code, 0) + 1
            first_occur = code_seen[code] == 1
            role = sval(ws.cell(row=r, column=c["记录表角色"]).value)
            site = sval(ws.cell(row=r, column=c["记录表站点"]).value)
            if first_occur and classify == "客观问题" and qa_type == "执行问题" \
                    and role == "质检" and site == "武汉库":
                error_person = inspector
            else:
                error_person = ""
            result_action = sval(ws.cell(row=r, column=c["记录表结果"]).value) or \
                            sval(ws.cell(row=r, column=c["记录表判定方式"]).value)
            imgs = []
            for _col, img in sorted(img_by_row.get(r, [])):
                fname = f"four_{r}_{len(imgs) + 1}.jpg"
                _write_optimized_image(img._data(), os.path.join(images_dir, fname))
                imgs.append(fname)
            records.append({
                "date": d17.strftime("%Y-%m-%d"),
                "code": code,
                "imgs": imgs,
                "cat": sval(ws.cell(row=r, column=c["记录表品类"]).value),
                "brand": sval(ws.cell(row=r, column=c["记录表品牌"]).value),
                "model": sval(ws.cell(row=r, column=c["记录表型号"]).value),
                "inspector": inspector,
                "resultAction": result_action,
                "isExecError": "执行问题" if error_person else "",
                "errorPerson": error_person,
                "level1": sval(ws.cell(row=r, column=c["记录表一级项"]).value),
                "level2": sval(ws.cell(row=r, column=c["记录表二级项"]).value),
                "detail": sval(ws.cell(row=r, column=c["记录表明细"]).value),
                "classify": classify,
                "problemType": sval(ws.cell(row=r, column=c["问题类型"]).value),
                "qaType": qa_type,
            })
    records.sort(key=lambda d: (d["date"], d["code"]))
    return daily, daily_cat, person_cat, records


def parse_qa(ws, phone_person_sum, phone_error_count, phone_diffs, mb_records, person_cat, fourcat_records):
    hdr = find_header_row(ws, ["质检人", "抽检量", "品类"])
    if hdr is None:
        raise SystemExit(f"「{ws.title}」表头未识别，请检查工作表结构")
    c_grp = col_of(ws, hdr, "小组")
    c_name = col_of(ws, hdr, "质检人")
    grp_cat = [
        ("【手机】", "📱"), ("笔记本质检", "💻"), ("平板电脑", "📱"),
        ("手表质检", "⌚️"), ("耳机质检", "🎧")]
    four_cat_map = {"笔记本质检": "笔记本", "平板电脑": "平板电脑",
                    "手表质检": "智能手表", "耳机质检": "耳机/耳麦"}

    people = []
    for r in range(hdr + 1, ws.max_row + 1):
        name = sval(ws.cell(row=r, column=c_name).value)
        if not name:
            continue
        grp = sval(ws.cell(row=r, column=c_grp).value)
        cat = next((e for k, e in grp_cat if grp.startswith(k)), "📱")
        people.append({"cat": cat, "grp": grp, "inspector": name})

    mb_penalty = {}
    for rec in mb_records:
        if rec["penalty"] == "是" and rec["inspector"]:
            mb_penalty[rec["inspector"]] = mb_penalty.get(rec["inspector"], 0) + 1

    fourcat_err = {}
    fourcat_total = {}
    for rec in fourcat_records:
        if rec["errorPerson"]:
            fourcat_err.setdefault((rec["errorPerson"], rec["cat"]), 0)
            fourcat_err[(rec["errorPerson"], rec["cat"])] += 1
        if rec["inspector"]:
            fourcat_total.setdefault((rec["inspector"], rec["cat"]), 0)
            fourcat_total[(rec["inspector"], rec["cat"])] += 1
    phone_total = {}
    for d in phone_diffs:
        if d["origInspector"]:
            phone_total[d["origInspector"]] = phone_total.get(d["origInspector"], 0) + 1

    out = []
    for p in people:
        name, grp = p["inspector"], p["grp"]
        if grp.startswith("【手机】"):
            samples = int(phone_person_sum.get(name, 0))
            errors = phone_error_count.get(name, 0)
            total = phone_total.get(name, 0)
            qa_miss = mb_penalty.get(name, 0)
        else:
            fc = four_cat_map.get(grp, "")
            samples = int(person_cat.get((name, fc), 0))
            errors = fourcat_err.get((name, fc), 0)
            total = fourcat_total.get((name, fc), 0)
            qa_miss = 0
        rate = round(errors / samples, 6) if samples else 0
        out.append({**p, "samples": samples, "errors": errors, "totalErrors": total,
                    "rate": rate, "ranking": "", "qaMiss": qa_miss})

    # 复现看板 RANK 公式（升序，差异率最低 = 第1名）
    for grp in ("【手机】后验一段", "【手机】后验二段"):
        members = [p for p in out if p["grp"] == grp]
        for p in members:
            better = sum(1 for q in members if q["rate"] < p["rate"])
            p["ranking"] = str(better + 1)
    return out


def parse_followup(ws):
    hdr = find_header_row(ws, ["失误人", "质检码"])
    items, high_freq = [], []
    if hdr is None:
        return items, high_freq
    c = {n: col_of(ws, hdr, n) for n in ["失误人", "质检码", "失误项", "失误类型", "失误时间", "失误原因"]}
    for r in range(hdr + 1, ws.max_row + 1):
        person = sval(ws.cell(row=r, column=c["失误人"]).value)
        if person:
            tv = ws.cell(row=r, column=c["失误时间"]).value
            items.append({
                "person": person,
                "code": sval(ws.cell(row=r, column=c["质检码"]).value),
                "item": sval(ws.cell(row=r, column=c["失误项"]).value),
                "errorType": sval(ws.cell(row=r, column=c["失误类型"]).value),
                "time": tv.strftime("%Y-%m-%d %H:%M:%S") if isinstance(tv, datetime) else sval(tv),
                "reason": sval(ws.cell(row=r, column=c["失误原因"]).value),
            })
        rank = sval(ws.cell(row=r, column=8).value)
        if rank:
            high_freq.append({
                "rank": rank,
                "name": sval(ws.cell(row=r, column=9).value),
                "count": num(ws.cell(row=r, column=10).value),
            })
    return items, high_freq


def parse_t1(ws):
    weekly, monthly = [], []
    for r in range(2, ws.max_row + 1):
        rng = sval(ws.cell(row=r, column=1).value)
        if rng:
            weekly.append({
                "range": rng,
                "topPersons": sval(ws.cell(row=r, column=2).value),
                "diffs": sval(ws.cell(row=r, column=3).value),
                "topItem": sval(ws.cell(row=r, column=4).value),
                "note": sval(ws.cell(row=r, column=5).value),
            })
        month = sval(ws.cell(row=r, column=7).value)
        if month:
            monthly.append({
                "month": month,
                "summary": sval(ws.cell(row=r, column=8).value),
                "sec1Top": sval(ws.cell(row=r, column=9).value),
                "sec1Miss": sval(ws.cell(row=r, column=10).value),
                "sec1Rate": sval(ws.cell(row=r, column=11).value),
                "sec2Top": sval(ws.cell(row=r, column=12).value),
                "sec2Miss": sval(ws.cell(row=r, column=13).value),
                "sec2Rate": sval(ws.cell(row=r, column=14).value),
                "detail": sval(ws.cell(row=r, column=15).value),
            })
    return weekly, monthly


def parse_transit(ws):
    archives, weekly_rates = [], []
    for r in range(1, ws.max_row + 1):
        v2 = sval(ws.cell(row=r, column=2).value)
        if v2.startswith("📊"):
            archives.append(v2)
        v15 = sval(ws.cell(row=r, column=15).value)
        if v15 in ("📱手机", "💻笔记本"):
            weekly_rates.append({
                "category": v15,
                "weeks": [num(ws.cell(row=r, column=c).value) for c in range(16, 20)],
            })
    return archives, weekly_rates


def parse_cuoti(xlsx_bytes, images_dir):
    """解析「错题集」表格（独立文档），返回错题记录与聚合统计，并提取单元格内截图"""
    import openpyxl
    tmp = os.path.join(REPO, "_tmp_cuoti.xlsx")
    with open(tmp, "wb") as f:
        f.write(xlsx_bytes)
    try:
        wb = openpyxl.load_workbook(tmp, data_only=True)
    finally:
        os.remove(tmp)
    return _parse_cuoti_wb(wb, images_dir)


def _parse_cuoti_wb(wb, images_dir):
    import re
    ws = find_sheet(wb, "错题")
    hdr = find_header_row(ws, ["错题知识点", "错题内容"])
    c = {n: col_of(ws, hdr, n) for n in
         ["日期", "站点", "考试类型", "品类", "错题知识点", "错题内容", "质检人员", "正确答案/解析"]}

    img_by_row = {}
    for img in ws._images:
        a = img.anchor
        row = (a._from.row if hasattr(a, "_from") else a.from_.row) + 1
        col = a._from.col if hasattr(a, "_from") else a.from_.col
        img_by_row.setdefault(row, []).append((col, img))

    img_count = 0
    records = []
    for r in range(hdr + 1, ws.max_row + 1):
        dt = ws.cell(row=r, column=c["日期"]).value
        if not dt:
            continue
        persons = [p for p in re.split(r"[,，、;；]", sval(ws.cell(row=r, column=c["质检人员"]).value)) if p]
        imgs = []
        for col, img in sorted(img_by_row.get(r, [])):
            if col < 7:
                continue
            fname = f"cuoti_{r}_{len(imgs) + 1}.jpg"
            _write_optimized_image(img._data(), os.path.join(images_dir, fname))
            imgs.append(fname)
            img_count += 1
        records.append({
            "date": sval(dt),
            "station": sval(ws.cell(row=r, column=c["站点"]).value),
            "examType": sval(ws.cell(row=r, column=c["考试类型"]).value),
            "cat": sval(ws.cell(row=r, column=c["品类"]).value),
            "kp": sval(ws.cell(row=r, column=c["错题知识点"]).value),
            "content": sval(ws.cell(row=r, column=c["错题内容"]).value),
            "persons": persons,
            "answer": sval(ws.cell(row=r, column=c["正确答案/解析"]).value),
            "imgs": imgs,
        })
    if img_count:
        print(f"    错题集截图 {img_count} 张已提取压缩")
    if not records:
        return {}

    records.sort(key=lambda x: x["date"])
    kp_map, person_map, daily_map, cat_map, exam_map = {}, {}, {}, {}, {}
    for rec in records:
        for p in rec["persons"]:
            person_map.setdefault(p, {"count": 0, "kps": {}})
            person_map[p]["count"] += 1
            person_map[p]["kps"][rec["kp"]] = person_map[p]["kps"].get(rec["kp"], 0) + 1
        kp_map.setdefault(rec["kp"], {"count": 0, "persons": set(), "cats": set(), "lastDate": ""})
        kp_map[rec["kp"]]["count"] += 1
        kp_map[rec["kp"]]["persons"].update(rec["persons"])
        kp_map[rec["kp"]]["cats"].add(rec["cat"])
        kp_map[rec["kp"]]["lastDate"] = max(kp_map[rec["kp"]]["lastDate"], rec["date"])
        daily_map[rec["date"]] = daily_map.get(rec["date"], 0) + 1
        if rec["cat"]:
            cat_map[rec["cat"]] = cat_map.get(rec["cat"], 0) + 1
        if rec["examType"]:
            exam_map[rec["examType"]] = exam_map.get(rec["examType"], 0) + 1

    kp_stats = sorted(
        ({"kp": k, "count": v["count"], "persons": len(v["persons"]),
          "cats": "/".join(sorted(v["cats"])), "lastDate": v["lastDate"]}
         for k, v in kp_map.items()),
        key=lambda x: (-x["count"], -x["persons"], x["kp"]))
    person_stats = sorted(
        ({"person": k, "count": v["count"],
          "kps": sorted(v["kps"].items(), key=lambda t: (-t[1], t[0]))[:10]}
         for k, v in person_map.items()),
        key=lambda x: (-x["count"], x["person"]))
    dates = [r["date"] for r in records]
    return {
        "meta": {
            "count": len(records),
            "kpCount": len(kp_map),
            "personCount": len(person_map),
            "catCount": len(cat_map),
            "dateMin": min(dates),
            "dateMax": max(dates),
        },
        "records": records,
        "kpStats": kp_stats,
        "personStats": person_stats,
        "daily": [{"date": d, "count": daily_map[d]} for d in sorted(daily_map)],
        "catStats": [{"cat": k, "count": cat_map[k]}
                     for k in sorted(cat_map, key=lambda c: -cat_map[c])],
        "examStats": [{"exam": k, "count": exam_map[k]}
                      for k in sorted(exam_map, key=lambda e: -exam_map[e])],
    }


# ---------- 培训动作（留底表除错题集外的 5 个工作表） ----------

def parse_training(wb):
    """解析留底表：每日推题/一对一辅导/晨会/周考模拟考/周考月考分数，
    并汇总待辅导预警名单（推题未达标 + 模拟考不合格 + 最近分数低于90）"""
    import re

    _NOT_PERSON = {"全合格", "无", "无不合格", "全部合格", "全员合格", "合格", "无人员"}

    def names(v):
        return [p for p in re.split(r"[,，、;；]", sval(v))
                if p and p not in _NOT_PERSON]

    out = {}

    # 每日推题记录
    ws = find_sheet(wb, "推题", required=False)
    if ws:
        hdr = find_header_row(ws, ["参考人数", "达标人数"])
        if hdr:
            c = {n: col_of(ws, hdr, n) for n in
                 ["日期", "品类", "考试阶段", "达标线", "参考人数", "达标人数",
                  "不合格人员", "补推错题人数", "账号禁用人数", "当日高频错题"]}
            records, daily, cat_map = [], {}, {}
            for r in range(hdr + 1, ws.max_row + 1):
                if not sval(ws.cell(row=r, column=c["日期"]).value):
                    continue
                d = sval(ws.cell(row=r, column=c["日期"]).value)[:10]
                cat = sval(ws.cell(row=r, column=c["品类"]).value)
                total = num(ws.cell(row=r, column=c["参考人数"]).value)
                passed = num(ws.cell(row=r, column=c["达标人数"]).value)
                fails = names(ws.cell(row=r, column=c["不合格人员"]).value)
                records.append({
                    "date": d, "cat": cat,
                    "stage": sval(ws.cell(row=r, column=c["考试阶段"]).value),
                    "passLine": num(ws.cell(row=r, column=c["达标线"]).value),
                    "total": total, "passed": passed, "failNames": fails,
                    "rePush": num(ws.cell(row=r, column=c["补推错题人数"]).value),
                    "banned": num(ws.cell(row=r, column=c["账号禁用人数"]).value),
                    "freqKp": sval(ws.cell(row=r, column=c["当日高频错题"]).value),
                })
                daily.setdefault(d, [0, 0])
                daily[d][0] += total
                daily[d][1] += passed
                cat_map.setdefault(cat, [0, 0, 0])
                cat_map[cat][0] += total
                cat_map[cat][1] += passed
                cat_map[cat][2] += len(fails)
            fail_map = {}
            for rec in records:
                for p in rec["failNames"]:
                    fail_map.setdefault(p, {"count": 0, "dates": []})
                    fail_map[p]["count"] += 1
                    if rec["date"] not in fail_map[p]["dates"]:
                        fail_map[p]["dates"].append(rec["date"])
            out["tuili"] = {
                "records": records,
                "daily": [{"date": d, "total": v[0], "passed": v[1]}
                          for d, v in sorted(daily.items())],
                "catStats": [{"cat": k, "total": v[0], "passed": v[1], "fails": v[2]}
                             for k, v in sorted(cat_map.items(), key=lambda t: -t[1][0])],
                "fails": sorted(({"person": p, "count": v["count"], "dates": sorted(v["dates"])}
                                 for p, v in fail_map.items()),
                                key=lambda x: (-x["count"], x["person"])),
            }

    # 一对一辅导记录
    ws = find_sheet(wb, "辅导", required=False)
    if ws:
        hdr = find_header_row(ws, ["被辅导人", "辅导内容"])
        if hdr:
            c = {n: col_of(ws, hdr, n) for n in
                 ["辅导日期", "被辅导人", "辅导品类", "辅导内容", "辅导时长(分钟)", "辅导方式(现场/线上)", "效果评估"]}
            records, person_map = [], {}
            for r in range(hdr + 1, ws.max_row + 1):
                if not sval(ws.cell(row=r, column=c["辅导日期"]).value):
                    continue
                p = sval(ws.cell(row=r, column=c["被辅导人"]).value)
                if not p:
                    continue
                minutes = num(ws.cell(row=r, column=c["辅导时长(分钟)"]).value)
                records.append({
                    "date": sval(ws.cell(row=r, column=c["辅导日期"]).value)[:10],
                    "person": p,
                    "cat": sval(ws.cell(row=r, column=c["辅导品类"]).value),
                    "content": sval(ws.cell(row=r, column=c["辅导内容"]).value),
                    "minutes": minutes,
                    "mode": sval(ws.cell(row=r, column=c["辅导方式(现场/线上)"]).value),
                    "effect": sval(ws.cell(row=r, column=c["效果评估"]).value),
                })
                person_map.setdefault(p, {"count": 0, "minutes": 0})
                person_map[p]["count"] += 1
                person_map[p]["minutes"] += minutes
            out["coach"] = {
                "records": records,
                "personStats": sorted(({"person": p, "count": v["count"], "minutes": v["minutes"]}
                                       for p, v in person_map.items()),
                                      key=lambda x: (-x["count"], x["person"])),
            }

    # 晨会讲解记录
    ws = find_sheet(wb, "晨会", required=False)
    if ws:
        hdr = find_header_row(ws, ["讲解主题", "参会人员"])
        if hdr:
            c = {n: col_of(ws, hdr, n) for n in ["日期", "讲解主题", "涉及品类", "讲解要点", "参会人员"]}
            records = []
            for r in range(hdr + 1, ws.max_row + 1):
                if not sval(ws.cell(row=r, column=c["日期"]).value):
                    continue
                records.append({
                    "date": sval(ws.cell(row=r, column=c["日期"]).value)[:10],
                    "topic": sval(ws.cell(row=r, column=c["讲解主题"]).value),
                    "cat": sval(ws.cell(row=r, column=c["涉及品类"]).value),
                    "points": sval(ws.cell(row=r, column=c["讲解要点"]).value),
                    "attendees": sval(ws.cell(row=r, column=c["参会人员"]).value),
                })
            out["meeting"] = {"records": records}

    # 周考模拟考试记录
    ws = find_sheet(wb, "模拟考试", required=False)
    if ws:
        hdr = find_header_row(ws, ["参考人数", "合格人数"])
        if hdr:
            c = {n: col_of(ws, hdr, n) for n in
                 ["考试日期", "品类", "考试阶段", "参考人数", "合格人数", "合格率", "最高分", "最低分", "不合格人员"]}
            records = []
            for r in range(hdr + 1, ws.max_row + 1):
                if not sval(ws.cell(row=r, column=c["考试日期"]).value):
                    continue
                total = num(ws.cell(row=r, column=c["参考人数"]).value)
                passed = num(ws.cell(row=r, column=c["合格人数"]).value)
                rate = num(ws.cell(row=r, column=c["合格率"]).value)
                rate = round(rate * 100, 1) if 0 < rate <= 1 \
                    else round(passed / total * 100, 1) if total else 0
                records.append({
                    "date": sval(ws.cell(row=r, column=c["考试日期"]).value)[:10],
                    "cat": sval(ws.cell(row=r, column=c["品类"]).value),
                    "stage": sval(ws.cell(row=r, column=c["考试阶段"]).value),
                    "total": total, "passed": passed, "rate": rate,
                    "maxScore": sval(ws.cell(row=r, column=c["最高分"]).value),
                    "minScore": sval(ws.cell(row=r, column=c["最低分"]).value),
                    "failNames": names(ws.cell(row=r, column=c["不合格人员"]).value),
                })
            fail_map = {}
            for rec in records:
                for p in rec["failNames"]:
                    fail_map.setdefault(p, {"count": 0, "dates": []})
                    fail_map[p]["count"] += 1
                    if rec["date"] not in fail_map[p]["dates"]:
                        fail_map[p]["dates"].append(rec["date"])
            out["exam"] = {
                "records": records,
                "fails": sorted(({"person": p, "count": v["count"], "dates": sorted(v["dates"])}
                                 for p, v in fail_map.items()),
                                key=lambda x: (-x["count"], x["person"])),
            }

    # 周考月考分数统计
    ws = find_sheet(wb, "分数统计", required=False)
    if ws:
        hdr = find_header_row(ws, ["质检人员", "分数"])
        if hdr:
            c = {n: col_of(ws, hdr, n) for n in ["时间", "质检人员", "周考/月考", "品类", "分数"]}
            records, person_map = [], {}
            for r in range(hdr + 1, ws.max_row + 1):
                if not sval(ws.cell(row=r, column=c["时间"]).value):
                    continue
                p = sval(ws.cell(row=r, column=c["质检人员"]).value)
                if not p:
                    continue
                rec = {
                    "date": sval(ws.cell(row=r, column=c["时间"]).value)[:10],
                    "person": p,
                    "examType": sval(ws.cell(row=r, column=c["周考/月考"]).value),
                    "cat": sval(ws.cell(row=r, column=c["品类"]).value),
                    "score": num(ws.cell(row=r, column=c["分数"]).value),
                }
                records.append(rec)
                person_map.setdefault(p, {"week": [], "month": []})
                key = "month" if "月" in rec["examType"] else "week"
                person_map[p][key].append({"date": rec["date"], "score": rec["score"]})
            people = []
            for p, v in person_map.items():
                for k in ("week", "month"):
                    v[k].sort(key=lambda x: x["date"])
                all_sorted = sorted(v["week"] + v["month"], key=lambda x: x["date"])
                first = all_sorted[0]["score"] if all_sorted else 0
                last = all_sorted[-1]["score"] if all_sorted else 0
                people.append({
                    "person": p,
                    "week": v["week"], "month": v["month"], "all": all_sorted,
                    "last": all_sorted[-1]["score"] if all_sorted else None,
                    "first": first,
                    "trend": round((last - first) / first * 100, 1) if first else 0,
                })
            out["scores"] = {
                "records": records,
                "people": sorted(people, key=lambda x: x["person"]),
            }

    # 待辅导预警汇总
    alert_map = {}

    def add_alert(person, reason, d):
        if not person:
            return
        a = alert_map.setdefault(person, {"reasons": [], "lastDate": ""})
        if reason not in a["reasons"]:
            a["reasons"].append(reason)
        if d > a["lastDate"]:
            a["lastDate"] = d

    for f in out.get("tuili", {}).get("fails", []):
        add_alert(f["person"], f"推题未达标{f['count']}次", f["dates"][-1] if f["dates"] else "")
    for f in out.get("exam", {}).get("fails", []):
        add_alert(f["person"], f"模拟考不合格{f['count']}次", f["dates"][-1] if f["dates"] else "")
    for p in out.get("scores", {}).get("people", []):
        if p["last"] is not None and p["last"] < 90 and p["all"]:
            add_alert(p["person"], f"最近分数{p['last']}", p["all"][-1]["date"])
    alerts = [{"person": p, "reasons": v["reasons"], "lastDate": v["lastDate"]}
              for p, v in alert_map.items()]
    alerts.sort(key=lambda x: x["person"])
    alerts.sort(key=lambda x: x["lastDate"] or "", reverse=True)
    out["alerts"] = alerts
    return out


# ---------- 飞书消息 ----------

FEISHU_OPEN_ID = "ou_d80917787fd37a300ccf1ef12ef58c40"


def send_feishu_message(text):
    """通过飞书 IM 给宗飞飞发送看板更新摘要（应用身份 + open_id）"""
    app_id, app_secret = get_app_credentials()
    r = _http_json(f"{BASE}/open-apis/auth/v3/tenant_access_token/internal", "POST",
                   body={"app_id": app_id, "app_secret": app_secret})
    if r.get("code") != 0:
        print(f"    警告: 获取 token 失败，跳过飞书消息（{r.get('msg')}）")
        return
    r = _http_json(f"{BASE}/open-apis/im/v1/messages?receive_id_type=open_id", "POST",
                   headers={"Authorization": f"Bearer {r['tenant_access_token']}"},
                   body={"receive_id": FEISHU_OPEN_ID, "msg_type": "text",
                         "content": json.dumps({"text": text})})
    if r.get("code") == 0:
        print("    已发送飞书消息 ✓")
    else:
        print(f"    警告: 飞书消息发送失败（{r.get('msg')}）")


def build_summary_message(d):
    m = d["meta"]
    lines = [f"【质量看板已更新】{m['monthLabel']}（数据至 {m['dateMax']}）"]
    lines.append(f"手机差异 {len(d['diffs'])} 条 | 四品类 {len(d['fourcatDiffs'])} 条 | 主板 {len(d['mb'])} 条")
    exec_map = {}
    for x in d["diffs"]:
        if x.get("qaJudgment") == "执行问题" and x.get("errorPerson"):
            exec_map[x["errorPerson"]] = exec_map.get(x["errorPerson"], 0) + 1
    for x in d["fourcatDiffs"]:
        if x.get("isExecError") and x.get("errorPerson"):
            exec_map[x["errorPerson"]] = exec_map.get(x["errorPerson"], 0) + 1
    top = sorted(exec_map.items(), key=lambda t: -t[1])[:3]
    if top:
        lines.append("执行问题TOP: " + "、".join(f"{p}{n}次" for p, n in top))
    cm = (d.get("cuoti") or {}).get("meta") or {}
    if cm.get("count"):
        top_kp = (d.get("cuoti").get("kpStats") or [{}])[0]
        lines.append(f"错题 {cm['count']} 条 · 高频知识点: {top_kp.get('kp', '')} {top_kp.get('count', 0)}次")
    alerts = (d.get("training") or {}).get("alerts") or []
    if alerts:
        shown = "、".join(f"{a['person']}({a['lastDate'][5:]})" for a in alerts[:5])
        lines.append(f"待辅导预警 {len(alerts)} 人: {shown}")
    lines.append("看板: https://zffzms1995.github.io/wuhan-quality-dashboard/")
    return "\n".join(lines)


# ---------- 主流程 ----------

def build_data(xlsx_bytes, images_dir):
    import openpyxl
    tmp = os.path.join(images_dir, "..", "_tmp_src.xlsx")
    with open(tmp, "wb") as f:
        f.write(xlsx_bytes)
    try:
        wb = openpyxl.load_workbook(tmp, data_only=True)
    finally:
        os.remove(tmp)

    ws_phone = find_sheet(wb, "手机品类")
    ws_four = find_sheet(wb, "四品类")
    ws_mb = find_sheet(wb, "主板审核")
    ws_qa = find_sheet(wb, "达成情况看板")
    ws_follow = find_sheet(wb, "工作表1", required=False)
    ws_t1 = find_sheet(wb, "周期Top", required=False)
    ws_transit = find_sheet(wb, "中转站", required=False)

    print("⑤ 解析主板审核 ...")
    mb, mb_img = parse_mb(ws_mb, images_dir)
    print("⑥ 解析手机品类 ...")
    phone_trend, phone_person_sum, phone_error_count, diffs = parse_phone(ws_phone, images_dir)
    print("⑦ 解析四品类 ...")
    four_daily, four_daily_cat, four_person_cat, four_records = parse_fourcat(ws_four, images_dir)
    print("⑧ 计算人员质量达成（复现表格公式） ...")
    qa = parse_qa(ws_qa, phone_person_sum, phone_error_count, diffs, mb,
                  four_person_cat, four_records)

    error_items, high_freq = ([], [])
    t1_weekly, t1_monthly = ([], [])
    archives, weekly_rates = ([], [])
    if ws_follow:
        error_items, high_freq = parse_followup(ws_follow)
    if ws_t1:
        t1_weekly, t1_monthly = parse_t1(ws_t1)
    if ws_transit:
        archives, weekly_rates = parse_transit(ws_transit)

    # T-1 质量数据（复现看板公式，T-1 = 昨天）
    yesterday = (date.today() - timedelta(days=1)).strftime("%Y-%m-%d")
    t1_rows = []
    t1_phone = {"category": "📱", "samples": 0, "objDiffs": 0, "execDiffs": 0, "execRate": 0}
    hdr_p = find_header_row(ws_phone, ["抽检日期", "总抽检量"])
    c_date = col_of(ws_phone, hdr_p, "抽检日期")
    c_samp = col_of(ws_phone, hdr_p, "总抽检量")
    c_obj = col_of(ws_phone, hdr_p, "总客观差异量")
    for r in range(hdr_p + 1, ws_phone.max_row + 1):
        v = ws_phone.cell(row=r, column=c_date).value
        if isinstance(v, datetime) and v.strftime("%Y-%m-%d") == yesterday:
            t1_phone["samples"] = num(ws_phone.cell(row=r, column=c_samp).value)
            t1_phone["objDiffs"] = num(ws_phone.cell(row=r, column=c_obj).value)
            break
    t1_phone["execDiffs"] = sum(1 for d in diffs
                                if d["date"] == yesterday and d["qaJudgment"] == "执行问题")
    t1_phone["execRate"] = round(t1_phone["execDiffs"] / t1_phone["samples"], 6) \
        if t1_phone["samples"] else 0
    t1_rows.append(t1_phone)
    t1_nb = {"category": "💻", "samples": 0, "objDiffs": 0, "execDiffs": 0, "execRate": 0}
    t1_nb["samples"] = sum(v[0] for d, v in four_daily_cat.get("笔记本", {}).items()
                           if d == yesterday)
    t1_nb["execDiffs"] = sum(1 for d in four_records
                             if d["date"] == yesterday and d["qaType"] == "执行问题")
    t1_nb["execRate"] = round(t1_nb["execDiffs"] / t1_nb["samples"], 6) \
        if t1_nb["samples"] else 0
    t1_rows.append(t1_nb)

    # meta
    all_dates = [d["date"] for d in diffs + four_records + mb if d["date"]]
    if all_dates:
        dmax = max(all_dates)
        y, m = int(dmax[:4]), int(dmax[5:7])
        month_label = f"{y}年{m}月"
        date_min, date_max = f"{y:04d}-{m:02d}-01", \
            f"{y:04d}-{m:02d}-{calendar.monthrange(y, m)[1]:02d}"
    else:
        month_label = date.today().strftime("%Y年%m月")
        date_min = date_max = date.today().strftime("%Y-%m-%d")

    dashboard = {
        "meta": {
            "monthLabel": month_label,
            "dateMin": date_min,
            "dateMax": date_max,
            "updatedAt": date.today().isoformat(),
        },
        "mb": mb,
        "diffs": diffs,
        "fourcatDiffs": four_records,
        "qaData": qa,
        "errorFollowup": {"items": error_items, "highFreq": high_freq},
        "t1Data": {"weekly": t1_weekly, "monthly": t1_monthly},
        "reviewData": {"t1": t1_rows, "weeklyRates": weekly_rates, "archives": archives},
    }
    mmdd = lambda s: s[5:].replace("-", "/")
    trend = {
        "phone": phone_trend,
        "fourCat": {cat: [{"date": mmdd(d), "samples": v[0], "diffs": v[1]}
                          for d, v in sorted(m.items())]
                    for cat, m in sorted(four_daily_cat.items())},
        "fourCatAgg": [{"date": mmdd(d), "samples": v[0], "diffs": v[1]}
                       for d, v in sorted(four_daily.items())],
    }
    return dashboard, trend, mb_img


def archive_current(images_dir, quality_token, cuoti_token):
    """把当前 dashboard_data.json + trend_data.json + images/ 归档到 archive/YYYY-MM/。

    必须在清空 images/ 之前调用（旧月份图片靠这一步留存）。
    每次运行幂等：数据无变化时跳过，不会重复提交大文件。
    """
    dd_path = os.path.join(REPO, "dashboard_data.json")
    td_path = os.path.join(REPO, "trend_data.json")
    if not (os.path.exists(dd_path) and os.path.exists(td_path)):
        return None
    with open(dd_path, encoding="utf-8") as f:
        dashboard = json.load(f)
    with open(td_path, encoding="utf-8") as f:
        trend = json.load(f)
    m = re.match(r"(\d{4})年(\d{1,2})月", dashboard.get("meta", {}).get("monthLabel", ""))
    if not m:
        return None
    month_key = f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    arch = os.path.join(ARCHIVE_DIR, month_key)
    os.makedirs(os.path.join(arch, "images"), exist_ok=True)
    # 归档数据里补上当时的飞书表格链接（首次归档时用当前运行的表 token 补齐）
    meta = dashboard.setdefault("meta", {})
    meta.setdefault("editUrl", f"{FEISHU_DOC_BASE}/sheets/{quality_token}")
    meta.setdefault("cuotiUrl", f"{FEISHU_DOC_BASE}/sheets/{cuoti_token}")
    data_str = json.dumps({"dashboard": dashboard, "trend": trend}, ensure_ascii=False, indent=2)
    data_path = os.path.join(arch, "data.json")
    if os.path.exists(data_path):
        with open(data_path, encoding="utf-8") as f:
            if f.read() == data_str:
                print(f"① 归档 {month_key} 无变化，跳过")
                return month_key
    with open(data_path, "w", encoding="utf-8") as f:
        f.write(data_str)
    arch_img = os.path.join(arch, "images")
    copied = 0
    for fn in os.listdir(images_dir):
        src = os.path.join(images_dir, fn)
        if os.path.isfile(src) and fn.lower().endswith((".png", ".jpg", ".jpeg")):
            shutil.copy2(src, os.path.join(arch_img, fn))
            copied += 1
    print(f"① 已归档: archive/{month_key}/（data.json + {copied} 张图片）")
    return month_key


def write_months_json(cur_month_key):
    """生成 months.json：看板月份下拉的选项列表（当前月标记 live）。"""
    months = {}
    if os.path.isdir(ARCHIVE_DIR):
        for name in os.listdir(ARCHIVE_DIR):
            if re.match(r"^\d{4}-\d{2}$", name) and \
                    os.path.exists(os.path.join(ARCHIVE_DIR, name, "data.json")):
                y, m = int(name[:4]), int(name[5:7])
                months[name] = {"month": name, "label": f"{y}年{m}月", "live": name == cur_month_key}
    if cur_month_key not in months:
        y, m = int(cur_month_key[:4]), int(cur_month_key[5:7])
        months[cur_month_key] = {"month": cur_month_key, "label": f"{y}年{m}月", "live": True}
    ordered = sorted(months.values(), key=lambda x: x["month"], reverse=True)
    with open(os.path.join(REPO, "months.json"), "w", encoding="utf-8") as f:
        json.dump(ordered, f, ensure_ascii=False, indent=2)
    return ordered


def main():
    ap = argparse.ArgumentParser(description="更新质量看板数据")
    ap.add_argument("--token", default=DEFAULT_FILE_TOKEN, help="飞书云表格文档 token")
    ap.add_argument("--cuoti-token", default=CUOTI_FILE_TOKEN, help="错题/培训留底表 token")
    ap.add_argument("--xlsx", help="本地 xlsx 文件路径（跳过飞书导出）")
    ap.add_argument("--no-commit", action="store_true", help="只生成数据文件，不提交 git")
    args = ap.parse_args()

    images_dir = os.path.join(REPO, "images")
    os.makedirs(images_dir, exist_ok=True)
    # 清图片之前先把当前数据归档，旧月份图片靠归档留存
    archive_current(images_dir, args.token, args.cuoti_token)
    # 清理旧月份图片，避免残留
    for fn in os.listdir(images_dir):
        if fn.lower().endswith((".png", ".jpg", ".jpeg")):
            os.remove(os.path.join(images_dir, fn))

    if args.xlsx:
        with open(args.xlsx, "rb") as f:
            xlsx_bytes = f.read()
        print(f"使用本地文件: {args.xlsx}")
    else:
        xlsx_bytes = fetch_xlsx(args.token)

    dashboard, trend, mb_img = build_data(xlsx_bytes, images_dir)

    print("⑨ 获取错题集与培训留底 ...")
    try:
        cuoti_bytes = fetch_xlsx(args.cuoti_token)
        import io as _io
        import openpyxl as _xl
        wb_c = _xl.load_workbook(_io.BytesIO(cuoti_bytes), data_only=True)
        dashboard["cuoti"] = _parse_cuoti_wb(wb_c, images_dir)
        dashboard["training"] = parse_training(wb_c)
        cm = dashboard["cuoti"].get("meta")
        if cm:
            print(f"    错题 {cm['count']} 条 / 知识点 {cm['kpCount']} 个 / "
                  f"涉及 {cm['personCount']} 人（{cm['dateMin']} ~ {cm['dateMax']}）")
        else:
            print("    提示: 错题集暂为空")
        tr = dashboard["training"]
        print(f"    培训动作: 推题 {len(tr.get('tuili', {}).get('records', []))} 条 / "
              f"辅导 {len(tr.get('coach', {}).get('records', []))} 次 / "
              f"晨会 {len(tr.get('meeting', {}).get('records', []))} 次 / "
              f"模拟考 {len(tr.get('exam', {}).get('records', []))} 场 / "
              f"分数 {len(tr.get('scores', {}).get('records', []))} 条 / "
              f"待辅导预警 {len(tr.get('alerts', []))} 人")
    except SystemExit as e:
        print(f"    警告: 留底表获取失败（{e}），本次跳过错题与培训数据")
        dashboard["cuoti"] = {}
        dashboard["training"] = {}

    # 留底表月份与主表不一致时提醒（换月时容易忘记切换留底表）
    cur_key = f"{dashboard['meta']['monthLabel'][:4]}-{int(dashboard['meta']['monthLabel'][5:-1]):02d}"
    cm = dashboard["cuoti"].get("meta") if isinstance(dashboard.get("cuoti"), dict) else None
    if cm and cm.get("dateMax") and cm["dateMax"][:7] != cur_key:
        print(f"    ⚠ 注意: 留底表数据到 {cm['dateMax']}，与主表月份 {cur_key} 不一致，"
              f"请确认是否该换新留底表（--cuoti-token）")

    # 记录当前数据对应的飞书表格链接（看板「去飞书改」按月份跳对表格）
    dashboard["meta"]["editUrl"] = f"{FEISHU_DOC_BASE}/sheets/{args.token}"
    dashboard["meta"]["cuotiUrl"] = f"{FEISHU_DOC_BASE}/sheets/{args.cuoti_token}"

    with open(os.path.join(REPO, "dashboard_data.json"), "w", encoding="utf-8") as f:
        json.dump(dashboard, f, ensure_ascii=False, indent=2)
    with open(os.path.join(REPO, "trend_data.json"), "w", encoding="utf-8") as f:
        json.dump(trend, f, ensure_ascii=False, indent=2)
    write_months_json(cur_key)

    print("\n========== 生成结果 ==========")
    print(f"主板审核: {len(dashboard['mb'])} 条（图片 {mb_img} 张）")
    print(f"手机差异: {len(dashboard['diffs'])} 条")
    print(f"四品类差异: {len(dashboard['fourcatDiffs'])} 条")
    print(f"QA人员: {len(dashboard['qaData'])} 人")
    print(f"手机每日趋势: {len(trend['phone'])} 天")
    print(f"月份: {dashboard['meta']['monthLabel']}"
          f"（{dashboard['meta']['dateMin']} ~ {dashboard['meta']['dateMax']}）")
    print("已写入: dashboard_data.json / trend_data.json / images/")

    if args.no_commit:
        return
    print("\n提交 git ...")
    r = subprocess.run(["git", "-C", REPO, "add", "-A"], capture_output=True, text=True)
    if r.returncode != 0:
        print(f"提示: git add 失败（{r.stderr.strip()}），该目录可能还不是 git 仓库，跳过提交")
        return
    r = subprocess.run(["git", "-C", REPO, "diff", "--cached", "--quiet"])
    if r.returncode == 0:
        print("数据无变化，跳过提交")
        return
    subprocess.run(["git", "-C", REPO, "commit", "-m",
                    f"数据更新: {dashboard['meta']['monthLabel']} (至 {dashboard['meta']['dateMax']})"],
                   capture_output=True, text=True)
    r = subprocess.run(["git", "-C", REPO, "remote", "get-url", "origin"],
                       capture_output=True, text=True)
    if r.returncode == 0:
        pushed = False
        for attempt in range(1, 4):
            p = subprocess.run(["git", "-C", REPO, "push"], capture_output=True, text=True)
            if p.returncode == 0:
                pushed = True
                break
            print(f"    push 失败（第{attempt}次）: {p.stderr.strip()[:80]}")
            time.sleep(5)
            p = subprocess.run(["git", "-C", REPO, "-c", "http.version=HTTP/1.1", "push"],
                               capture_output=True, text=True)
            if p.returncode == 0:
                pushed = True
                break
            print(f"    push 失败（HTTP/1.1）: {p.stderr.strip()[:80]}")
            time.sleep(5)
        if not pushed:
            raise SystemExit("git push 多次失败（GitHub 连接不稳定），请稍后手动执行 git push")
        print("已推送到 GitHub Pages ✓")
        print("发送飞书消息 ...")
        send_feishu_message(build_summary_message(dashboard))
    else:
        print("提示: 尚未配置远程仓库（origin），本地已提交。配置后执行 git push 即可发布。")


if __name__ == "__main__":
    main()
